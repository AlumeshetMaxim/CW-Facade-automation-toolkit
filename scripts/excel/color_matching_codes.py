"""Color Excel cells that match a list of item codes.

Public template for schedule checking workflows.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

INPUT_XLSX = Path("./examples/schedule.xlsx")
OUTPUT_XLSX = Path("./reports/schedule_colored.xlsx")
CODES = ["RC2K", "TC2F", "P27A"]
FILL = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")


def color_matching_cells(input_xlsx: Path, output_xlsx: Path, codes: list[str]) -> int:
    workbook = load_workbook(input_xlsx)
    codes_norm = {code.strip().upper() for code in codes}
    count = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if str(cell.value or "").strip().upper() in codes_norm:
                    cell.fill = FILL
                    count += 1
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)
    return count


def main() -> None:
    count = color_matching_cells(INPUT_XLSX, OUTPUT_XLSX, CODES)
    print(f"Colored cells: {count}. Output: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
